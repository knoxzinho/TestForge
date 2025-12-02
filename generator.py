# generator.py  -- Versão revisada completa
# Referência do arquivo original enviado pelo usuário: :contentReference[oaicite:1]{index=1}

import os
import sys
import json
import re
import datetime
import docx
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ======================
# CONFIGURAÇÃO
# ======================

API_KEY = "API_KEY"
client = genai.Client(api_key=API_KEY)

PASTA_DOCS = "Documentações"
EXCEL_OUTPUT = "cenarios_de_teste.xlsx"
DEBUG = False
PROMPT_MAX_CHARS = 15000

# Lista completa de categorias conforme prompt atualizado
TODAS_AS_LISTAS = [
    "cenarios_funcionais",
    "cenarios_negativos",
    "cenarios_borda",
    "cenarios_integracao",
    "cenarios_usabilidade",
    "cenarios_carga",
    "cenarios_estresse",
    "cenarios_aceitacao",
    "cenarios_smoke",
    "cenarios_exploratorios",
    "cenarios_compatibilidade",
    "cenarios_recuperacao",
    "cenarios_seguranca"
]

print("🚀 TestForge (revisado) iniciado.")

# =============================================
# 1) EXTRAÇÃO DOS REQUISITOS DE UM DOCX
# =============================================

def extrair_requisitos_docx(caminho):
    """Extrai seções do docx (títulos em negrito são seções)."""
    doc = docx.Document(caminho)
    sections = []
    sec_atual = {"title": "", "requirements": []}

    for para in doc.paragraphs:
        texto = para.text.strip()
        if not texto:
            continue

        # Títulos detectados por runs em negrito
        if para.runs and any(r.bold for r in para.runs):
            if sec_atual["title"]:
                sections.append(sec_atual)
            sec_atual = {"title": texto, "requirements": []}
        else:
            sec_atual["requirements"].append({"text": texto})

    # adicionar última seção
    if sec_atual["title"]:
        sections.append(sec_atual)

    return sections

# =================================================================================================
# 2) PROMPT OTIMIZADO PARA O GEMINI (mantém campos do prompt real)
# =================================================================================================

QA_PROMPT = """
Você é um Engenheiro de QA Sênior e Especialista em Automação com 15+ anos de experiência.

Sua missão: Analisar a funcionalidade e gerar testes completos + análise preditiva de bugs.

NÃO produza explicações fora do JSON. Apenas JSON válido em português (PT-BR).

ESTRUTURA DO JSON FINAL (mantenha todas as chaves mesmo que vazias):
{
  "meta_info": {
    "funcionalidade_alvo": "",
    "data_geracao": "",
    "complexidade_percebida": ""
  },
  "analise_preditiva_bugs": {
    "estimativa_total_bugs_esperados": "",
    "densidade_deifeitos_por_area": {
       "funcional": "",
       "seguranca": "",
       "usabilidade": "",
       "integracao": ""
    },
    "top_3_areas_risco_critico": [],
    "justificativa_analise": ""
  },
  "analise_requisitos": {
    "riscos_identificados": [],
    "suposicoes": []
  },

  "cenarios_funcionais": [],
  "cenarios_negativos": [],
  "cenarios_borda": [],
  "cenarios_integracao": [],
  "cenarios_usabilidade": [],
  "cenarios_carga": [],
  "cenarios_estresse": [],
  "cenarios_aceitacao": [],
  "cenarios_smoke": [],
  "cenarios_exploratorios": [],
  "cenarios_compatibilidade": [],
  "cenarios_recuperacao": [],
  "cenarios_seguranca": [],

  "metricas_qualidade": {
      "cobertura_caminhos_logicos": "",
      "prioridade_automacao": ""
  }
}

REGRAS:
- Se uma categoria não for aplicável, retorne [].
- JSON deve ser 100% válido e sem texto fora do bloco JSON.
"""

def build_prompt(sections, nome_funcionalidade=None):
    """Cria o prompt a enviar ao modelo."""
    combined = ""
    for section in sections:
        combined += f"\nSEÇÃO: {section['title']}\n"
        for req in section["requirements"]:
            linha = req["text"].strip().replace("\n", " ")
            combined += f"- {linha}\n"

    if nome_funcionalidade:
        header = f"FUNCIONALIDADE_ALVO: {nome_funcionalidade}\n"
    else:
        header = ""

    payload = QA_PROMPT + "\n\n" + header + "\nREQUISITOS_ANALISADOS:\n" + combined
    # garantir limite
    if len(payload) > PROMPT_MAX_CHARS:
        print("⚠️ Requisitos muito longos — compactando para evitar corte do modelo...")
        payload = payload[:PROMPT_MAX_CHARS]
        payload = payload.rsplit("\n", 1)[0] + "\n... (conteúdo reduzido automaticamente)"
    return payload

# ==================================
# 3) SANITIZADOR DE JSON (robusto)
# ==================================

def limpar_json_bruto(texto):
    """
    Extrai o primeiro JSON balanceado da saída do modelo.
    Estratégia: encontra primeiro '{' e fecha contando profundidade, respeitando strings e escapes.
    """
    if not texto or "{" not in texto:
        return texto

    start = texto.find("{")
    depth = 0
    in_string = False
    escape = False

    for i in range(start, len(texto)):
        ch = texto[i]
        if ch == '"' and not escape:
            in_string = not in_string
        if ch == "\\" and not escape:
            escape = True
            continue
        else:
            escape = False

        if not in_string:
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return texto[start:i+1]

    # fallback: de primeiro { até último }
    end = texto.rfind("}")
    if end != -1 and start < end:
        return texto[start:end+1]

    return texto

# ====================
# 4) CHAMADA AO MODELO
# ====================

def gerar_cenarios(prompt):
    """Chama o Gemini de forma segura; retorna texto bruto."""
    try:
        resp = client.models.generate_content(
            model="models/gemini-2.5-flash",
            contents=prompt
        )
        if DEBUG:
            print("🔍 DEBUG - resposta bruta do Gemini:\n", resp.text)
        return resp.text
    except Exception as e:
        print("❌ Erro ao chamar o modelo:", e)
        return ""

# =============================================
# 5) NORMALIZAR E VALIDAR O JSON
# =============================================

def validar_json(json_data):
    """Garante que todas as categorias existam e tenham tipo correto."""
    if not isinstance(json_data, dict):
        json_data = {}

    # garantir listas
    for categoria in TODAS_AS_LISTAS:
        if categoria not in json_data or not isinstance(json_data[categoria], list):
            json_data[categoria] = []

    # garantir blocos meta/analise/metricas
    if "meta_info" not in json_data or not isinstance(json_data["meta_info"], dict):
        json_data["meta_info"] = {}
    if "analise_preditiva_bugs" not in json_data or not isinstance(json_data["analise_preditiva_bugs"], dict):
        json_data["analise_preditiva_bugs"] = {}
    if "analise_requisitos" not in json_data or not isinstance(json_data["analise_requisitos"], dict):
        json_data["analise_requisitos"] = {}
    if "metricas_qualidade" not in json_data or not isinstance(json_data["metricas_qualidade"], dict):
        json_data["metricas_qualidade"] = {}

    return json_data

# =============================
# 6) UTILITÁRIAS PARA COERÇÃO
# =============================

def safe_to_list(value):
    """Garante que o valor seja uma lista de strings, mesmo que JSON venha inconsistente."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value]
    # se for string, quebrar por linhas se houver; senão encapsular
    if isinstance(value, str):
        lines = [l.strip() for l in value.splitlines() if l.strip()]
        return lines if lines else [value]
    # objeto qualquer -> transformar em string única
    return [str(value)]

def safe_get(tc, key):
    """Retorna campo coerente do cenário."""
    return tc.get(key, "")

# =============================================
# 7) ESCREVER ABA DO EXCEL (meta + cenários unidos)
# =============================================

def escrever_aba(ws, json_data):
    """
    Escreve no worksheet:
    - Cabeçalho com meta_info e analise_preditiva_bugs
    - Tabela única contendo todos os cenários das categorias (com coluna Tipo de Cenário)
    """
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", fill_type="solid")

    # --- topo: meta_info e analise_preditiva_bugs ---
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="Meta Info & Análise Preditiva (gerado automaticamente)").font = Font(bold=True, size=12)
    row += 1

    meta = json_data.get("meta_info", {})
    ws.cell(row=row, column=1, value="Funcionalidade Alvo:")
    ws.cell(row=row, column=2, value=meta.get("funcionalidade_alvo", ""))
    row += 1

    ws.cell(row=row, column=1, value="Data Geração:")
    ws.cell(row=row, column=2, value=meta.get("data_geracao", ""))
    ws.cell(row=row, column=3, value="Complexidade Percebida:")
    ws.cell(row=row, column=4, value=meta.get("complexidade_percebida", ""))
    row += 1

    analise = json_data.get("analise_preditiva_bugs", {})
    ws.cell(row=row, column=1, value="Estimativa total bugs esperados:")
    ws.cell(row=row, column=2, value=analise.get("estimativa_total_bugs_esperados", ""))
    row += 1

    ws.cell(row=row, column=1, value="Top 3 áreas de maior risco:")
    top3 = analise.get("top_3_areas_risco_critico", [])
    ws.cell(row=row, column=2, value=", ".join(top3) if isinstance(top3, list) else str(top3))
    row += 2

    # --- cabeçalho da tabela de cenários ---
    headers = ["ID", "Título", "Descrição", "Pré-condições", "Passos", "Resultado Esperado", "Tipo de Cenário"]
    start_table_row = row
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

    # --- juntar todas as categorias ---
    total_inseridos = 0
    for categoria in TODAS_AS_LISTAS:
        itens = json_data.get(categoria, [])
        if not isinstance(itens, list):
            # tentar coerção
            itens = [itens]

        for tc in itens:
            # coerções seguras
            id_ = safe_get(tc, "id") or safe_get(tc, "titulo")  # tenta algo decente se id ausente
            titulo = safe_get(tc, "titulo")
            descricao = safe_get(tc, "descricao")
            pre = safe_to_list(tc.get("pre_condicao", tc.get("pre_condicoes", [])))
            passos = safe_to_list(tc.get("passos", []))
            dados_teste = safe_get(tc, "dados_teste")
            resultado = safe_get(tc, "resultado_esperado")

            # inserir linha
            values = [
                id_,
                titulo,
                descricao,
                "\n".join(pre),
                "\n".join(passos),
                resultado,
                categoria
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
            total_inseridos += 1

    # Ajuste colunas
    col_widths = [15, 30, 55, 30, 45, 40, 25]
    for i, width in enumerate(col_widths, start=1):
        try:
            ws.column_dimensions[chr(64 + i)].width = width
        except Exception:
            pass

    # Congelar cabeçalho da tabela
    # freeze_panes em Excel: célula logo abaixo do header da tabela
    ws.freeze_panes = f"A{start_table_row + 1}"

    return total_inseridos

# ============================================================
# 8) PROCESSAR TODOS OS DOCX NA PASTA E GERAR EXCEL
# ============================================================

def sanitize_sheet_name(name, existing_names):
    """Garante nome de aba válido (<=31 chars) e único."""
    base = name[:31]
    candidate = base
    i = 1
    while candidate in existing_names:
        suffix = f"_{i}"
        allowed = 31 - len(suffix)
        candidate = base[:allowed] + suffix
        i += 1
    return candidate

if __name__ == "__main__":

    # checagens iniciais
    if not os.path.exists(PASTA_DOCS):
        print(f"❌ Pasta '{PASTA_DOCS}' não encontrada. Crie e coloque seus .docx lá.")
        sys.exit(1)

    arquivos = [f for f in os.listdir(PASTA_DOCS) if f.lower().endswith(".docx")]
    if not arquivos:
        print("❌ Nenhum documento .docx encontrado na pasta Documentações.")
        sys.exit(1)

    wb = Workbook()
    # remover sheet default
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    existing_sheet_names = set()
    summary_stats = {}

    for arquivo in sorted(arquivos):
        caminho = os.path.join(PASTA_DOCS, arquivo)
        nome_base = os.path.splitext(arquivo)[0]
        nome_aba = sanitize_sheet_name(nome_base, existing_sheet_names)
        existing_sheet_names.add(nome_aba)

        print(f"\n📄 Processando: {arquivo}")

        # extrair requisitos
        try:
            sections = extrair_requisitos_docx(caminho)
        except Exception as e:
            print(f"❌ Erro ao ler '{arquivo}': {e}")
            continue

        # montar prompt (inclui nome do documento como funcionalidade alvo)
        prompt = build_prompt(sections, nome_funcionalidade=nome_base)

        # chamar modelo
        resposta = gerar_cenarios(prompt)
        if not resposta:
            print("⚠️ Resposta vazia do modelo; pulando arquivo.")
            continue

        # salvar saída bruta para auditoria por arquivo
        raw_out_path = f"raw_{nome_base}.txt"
        try:
            with open(raw_out_path, "w", encoding="utf-8") as f:
                f.write(resposta)
        except Exception:
            pass

        # sanitizar JSON
        resposta_limpa = limpar_json_bruto(resposta)

        # tentar carregar JSON
        try:
            json_data = json.loads(resposta_limpa)
        except Exception as e:
            print("⚠️ JSON inválido — tentando extrair novamente e recarregar. Erro:", e)
            resposta_limpa2 = limpar_json_bruto(resposta_limpa)
            try:
                json_data = json.loads(resposta_limpa2)
            except Exception as e2:
                print("❌ Não foi possível interpretar JSON desse arquivo. Veja", raw_out_path)
                if DEBUG:
                    print("DEBUG - saída bruta:\n", resposta)
                continue

        # validar/normalizar
        json_data = validar_json(json_data)

        # preencher meta_info defaults se ausentes
        meta = json_data.get("meta_info", {})
        if "data_geracao" not in meta or not meta.get("data_geracao"):
            meta["data_geracao"] = datetime.date.today().isoformat()
        if "funcionalidade_alvo" not in meta or not meta.get("funcionalidade_alvo"):
            meta["funcionalidade_alvo"] = nome_base
        json_data["meta_info"] = meta

        # criar aba e escrever
        ws = wb.create_sheet(title=nome_aba)
        try:
            count = escrever_aba(ws, json_data)
            summary_stats[nome_aba] = count
            print(f"✅ Inseridos {count} cenários na aba '{nome_aba}'")
        except Exception as e:
            print(f"❌ Erro ao escrever aba '{nome_aba}': {e}")
            continue

    # adicionar aba de resumo (opcional) com contagem por documento
    try:
        ws_sum = wb.create_sheet(title="Resumo")
        ws_sum["A1"] = "Resumo de cenários por documento"
        ws_sum["A1"].font = Font(bold=True)
        row = 3
        ws_sum["A2"] = "Documento"
        ws_sum["B2"] = "Total Cenários"
        ws_sum["A2"].font = Font(bold=True)
        ws_sum["B2"].font = Font(bold=True)

        for nome, qtd in summary_stats.items():
            ws_sum.cell(row=row, column=1, value=nome)
            ws_sum.cell(row=row, column=2, value=qtd)
            row += 1
    except Exception:
        pass

    # salvar arquivo
    try:
        wb.save(EXCEL_OUTPUT)
        print(f"\n🎉 Finalizado! Excel gerado: {EXCEL_OUTPUT}")
    except Exception as e:
        print("❌ Erro ao salvar Excel:", e)
        sys.exit(1)