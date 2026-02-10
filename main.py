"""
TestForge - Gerador Inteligente de Plano de Testes Automatizado
Engenheiro de QA Sênior & Especialista em Automação

Objetivo: Ler documentação técnica e gerar casos de teste estruturados em Excel
com suporte a GitHub Copilot para geração inteligente de cenários.
"""

import os
import json
import logging
from pathlib import Path
from typing import List, Dict, Optional
import docx
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openai
from dotenv import load_dotenv

# ============================================================================
# CONFIGURAÇÕES E LOGGING
# ============================================================================

load_dotenv()

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('testforge.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configurar chave da API OpenAI
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
if not OPENAI_API_KEY:
    logger.warning("OPENAI_API_KEY não configurada. Funcionalidade de IA será desabilitada.")
    OPENAI_API_KEY = None
else:
    openai.api_key = OPENAI_API_KEY

# Configurações
DOCS_FOLDER = "Docs"
OUTPUT_FILE = "Plano_de_Testes.xlsx"
SUPPORTED_EXTENSIONS = {'.docx', '.pdf', '.txt'}


# ============================================================================
# EXTRAÇÃO DE TEXTO
# ============================================================================

class TextExtractor:
    """Classe responsável por extrair texto de diferentes formatos."""

    @staticmethod
    def extract_from_docx(file_path: str) -> str:
        """Extrai texto de arquivo .docx"""
        try:
            doc = docx.Document(file_path)
            text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            logger.info(f"Texto extraído de {file_path}")
            return text
        except Exception as e:
            logger.error(f"Erro ao extrair DOCX de {file_path}: {str(e)}")
            raise

    @staticmethod
    def extract_from_pdf(file_path: str) -> str:
        """Extrai texto de arquivo .pdf"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() + "\n"
            logger.info(f"Texto extraído de {file_path}")
            return text
        except Exception as e:
            logger.error(f"Erro ao extrair PDF de {file_path}: {str(e)}")
            raise

    @staticmethod
    def extract_from_txt(file_path: str) -> str:
        """Extrai texto de arquivo .txt"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
            logger.info(f"Texto extraído de {file_path}")
            return text
        except Exception as e:
            logger.error(f"Erro ao extrair TXT de {file_path}: {str(e)}")
            raise

    @staticmethod
    def extract_text(file_path: str) -> str:
        """Extrai texto baseado na extensão do arquivo."""
        extension = Path(file_path).suffix.lower()

        if extension == '.docx':
            return TextExtractor.extract_from_docx(file_path)
        elif extension == '.pdf':
            return TextExtractor.extract_from_pdf(file_path)
        elif extension == '.txt':
            return TextExtractor.extract_from_txt(file_path)
        else:
            raise ValueError(f"Formato não suportado: {extension}")


# ============================================================================
# GERAÇÃO INTELIGENTE DE CENÁRIOS COM IA
# ============================================================================

class TestCaseGenerator:
    """Classe responsável por gerar casos de teste usando GitHub Copilot (OpenAI)."""

    MODEL = "gpt-3.5-turbo"
    MAX_TOKENS = 2000

    @staticmethod
    def generate_test_cases(document_text: str, file_name: str) -> List[Dict]:
        """
        Gera casos de teste de forma inteligente usando OpenAI.

        Args:
            document_text: Texto extraído do documento
            file_name: Nome do arquivo para contexto

        Returns:
            Lista de dicionários com casos de teste
        """

        if not OPENAI_API_KEY:
            logger.warning("OpenAI API não configurada. Gerando casos de teste básicos.")
            return TestCaseGenerator._generate_basic_cases(document_text, file_name)

        prompt = TestCaseGenerator._build_prompt(document_text, file_name)

        try:
            response = openai.ChatCompletion.create(
                model=TestCaseGenerator.MODEL,
                messages=[
                    {
                        "role": "system",
                        "content": "Você é um Engenheiro de QA Sênior especialista em automação de testes. "
                                   "Gere casos de teste estruturados em formato JSON baseado na documentação fornecida."
                    },
                    {"role": "user", "content": prompt}
                ],
                max_tokens=TestCaseGenerator.MAX_TOKENS,
                temperature=0.7
            )

            response_text = response.choices[0].message.content
            
            # Tentar extrair JSON da resposta
            test_cases = TestCaseGenerator._parse_response(response_text)
            
            if test_cases:
                logger.info(f"Gerados {len(test_cases)} casos de teste para {file_name}")
                return test_cases
            else:
                logger.warning("Não foi possível parsear resposta da IA. Gerando casos básicos.")
                return TestCaseGenerator._generate_basic_cases(document_text, file_name)

        except Exception as e:
            logger.error(f"Erro ao chamar OpenAI API: {str(e)}")
            return TestCaseGenerator._generate_basic_cases(document_text, file_name)

    @staticmethod
    def _build_prompt(document_text: str, file_name: str) -> str:
        """Constrói o prompt para a IA gerar casos de teste."""

        # Limitar tamanho do texto para não exceder limites de tokens
        text_preview = document_text[:2000] if len(document_text) > 2000 else document_text

        return f"""\nAnalise a seguinte documentação técnica e gere casos de teste estruturados:\n\nARQUIVO: {file_name}\n\nDOCUMENTAÇÃO:\n{text_preview}\n\nGere EXATAMENTE 5 casos de teste no formato JSON com a seguinte estrutura:\n{{\n  \"test_cases\": [\n    {{\n      \"id\": \"TC001\",\n      \"titulo\": \"Título descritivo do cenário\",\n      \"precondições\": \"Pré-condições necessárias\",\n      \"passos\": \"1. Passo 1\n2. Passo 2\n3. Passo 3\",\n      \"resultado_esperado\": \"Descrição do resultado esperado\",\n      \"prioridade\": \"Alta\" (ou \"Média\" ou \"Baixa\")\n    }},\n    ...\n  ]\n}}\n\nRequisitos:\n- ID único para cada teste (TC001, TC002, etc)\n- Títulos claros e específicos\n- Pré-condições detalhadas\n- Passos numerados e objetivos\n- Resultado esperado verificável\n- Prioridade baseada em impacto funcional\n- Responda APENAS com JSON válido\n""" 

    @staticmethod
    def _parse_response(response_text: str) -> List[Dict]:
        """Faz parse da resposta JSON da IA."""

        try:
            # Tentar extrair JSON da resposta
            import re
            json_match = re.search(r'\{[\s\S]*\}', response_text)
            
            if json_match:
                json_str = json_match.group(0)
                data = json.loads(json_str)
                
                if isinstance(data, dict) and 'test_cases' in data:
                    return data['test_cases']
                elif isinstance(data, list):
                    return data

            return None

        except json.JSONDecodeError:
            logger.error("Erro ao fazer parse de JSON da resposta da IA")
            return None

    @staticmethod
    def _generate_basic_cases(document_text: str, file_name: str) -> List[Dict]:
        """Gera casos de teste básicos quando IA não está disponível.\n        Usa análise simples do texto para criar cenários."""

        test_cases = []
        
        # Análise simples para gerar casos
        lines = document_text.split('\n')
        relevant_lines = [line.strip() for line in lines if len(line.strip()) > 10][:5]

        for idx, line in enumerate(relevant_lines, 1):
            test_case = {
                "id": f"TC{idx:03d}",
                "titulo": f"Teste de Funcionalidade - {line[:50]}",
                "precondições": "Aplicação instalada e configurada",
                "passos": f"1. Acessar a funcionalidade\n2. Verificar: {line}\n3. Validar resultado",
                "resultado_esperado": "Funcionalidade funcionando conforme esperado",
                "prioridade": "Média" if idx % 3 == 0 else "Alta"
            }
            test_cases.append(test_case)

        if not test_cases:
            # Gerar caso padrão se documento está vazio
            test_cases = [{
                "id": "TC001",
                "titulo": "Teste Padrão - Validação do Documento",
                "precondições": "Documento disponível",
                "passos": "1. Ler documento\n2. Validar conteúdo\n3. Confirmar estrutura",
                "resultado_esperado": "Documento processado com sucesso",
                "prioridade": "Alta"
            }]

        logger.info(f"Gerados {len(test_cases)} casos de teste básicos para {file_name}")
        return test_cases


# ============================================================================
# GERAÇÃO DE ARQUIVO EXCEL
# ============================================================================

class ExcelGenerator:
    """Classe responsável por gerar o arquivo Excel com os casos de teste."""

    # Estilos
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    COLUMNS = [
        ("ID do Teste", 15),
        ("Título do Cenário", 35),
        ("Pré-condições", 30),
        ("Passos para Reprodução", 40),
        ("Resultado Esperado", 30),
        ("Prioridade", 12)
    ]

    @staticmethod
    def create_workbook(documents_data: Dict[str, List[Dict]]) -> str:
        """Cria o arquivo Excel com os casos de teste.

        Args:
            documents_data: Dicionário com {nome_arquivo: [casos_de_teste]}

        Returns:
            Caminho do arquivo criado
        """

        wb = Workbook()
        wb.remove(wb.active)  # Remove sheet padrão

        for file_name, test_cases in documents_data.items():
            sheet_name = ExcelGenerator._sanitize_sheet_name(file_name)
            ws = wb.create_sheet(title=sheet_name)

            # Criar cabeçalho
            for col_idx, (col_name, col_width) in enumerate(ExcelGenerator.COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = ExcelGenerator.HEADER_FILL
                cell.font = ExcelGenerator.HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = ExcelGenerator.BORDER
                ws.column_dimensions[chr(64 + col_idx)].width = col_width

            # Adicionar dados dos casos de teste
            for row_idx, test_case in enumerate(test_cases, 2):
                ws.cell(row=row_idx, column=1).value = test_case.get('id', '')
                ws.cell(row=row_idx, column=2).value = test_case.get('titulo', '')
                ws.cell(row=row_idx, column=3).value = test_case.get('precondições', '')
                ws.cell(row=row_idx, column=4).value = test_case.get('passos', '')
                ws.cell(row=row_idx, column=5).value = test_case.get('resultado_esperado', '')
                ws.cell(row=row_idx, column=6).value = test_case.get('prioridade', '')

                # Aplicar estilos
                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = ExcelGenerator.BORDER
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Congelar a linha de cabeçalho
            ws.freeze_panes = "A2"

        wb.save(OUTPUT_FILE)
        logger.info(f"Arquivo Excel criado: {OUTPUT_FILE}")
        return OUTPUT_FILE

    @staticmethod
    def _sanitize_sheet_name(file_name: str, max_length: int = 31) -> str:
        """Sanitiza o nome do arquivo para ser usado como nome de sheet.
        Excel limita nomes de sheet a 31 caracteres."""

        # Remove extensão
        name = Path(file_name).stem

        # Remove caracteres inválidos
        invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '_')

        # Limitar tamanho
        name = name[:max_length]

        return name if name else "Sheet1"


# ============================================================================
# CONTROLADOR PRINCIPAL
# ============================================================================

class TestForgeController:
    """Controlador principal da aplicação."""

    @staticmethod
    def validate_docs_folder() -> bool:
        """Valida se a pasta Docs existe e contém arquivos."""

        if not os.path.exists(DOCS_FOLDER):
            logger.error(f"Pasta '{DOCS_FOLDER}' não encontrada!")
            return False

        files = TestForgeController.get_supported_files()
        
        if not files:
            logger.error(f"Nenhum arquivo suportado ({SUPPORTED_EXTENSIONS}) encontrado em '{DOCS_FOLDER}'")
            return False

        logger.info(f"Encontrados {len(files)} arquivo(s) para processar")
        return True

    @staticmethod
    def get_supported_files() -> List[str]:
        """Retorna lista de arquivos suportados na pasta Docs."""

        if not os.path.exists(DOCS_FOLDER):
            return []

        files = []
        for file in os.listdir(DOCS_FOLDER):
            if Path(file).suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append(os.path.join(DOCS_FOLDER, file))

        return files

    @staticmethod
    def process_documents() -> bool:
        """Processa todos os documentos e gera o Excel."""

        try:
            if not TestForgeController.validate_docs_folder():
                return False

            documents_data = {}
            files = TestForgeController.get_supported_files()

            for file_path in files:
                file_name = os.path.basename(file_path)
                logger.info(f"Processando: {file_name}")

                try:
                    # Extrair texto
                    text = TextExtractor.extract_text(file_path)

                    if not text.strip():
                        logger.warning(f"Arquivo {file_name} está vazio!")
                        documents_data[file_name] = []
                        continue

                    # Gerar casos de teste
                    test_cases = TestCaseGenerator.generate_test_cases(text, file_name)
                    documents_data[file_name] = test_cases

                except Exception as e:
                    logger.error(f"Erro ao processar {file_name}: {str(e)}")
                    documents_data[file_name] = []

            # Gerar arquivo Excel
            if documents_data:
                ExcelGenerator.create_workbook(documents_data)
                logger.info("✓ Plano de testes gerado com sucesso!")
                return True
            else:
                logger.error("Nenhum documento foi processado com sucesso")
                return False

        except Exception as e:
            logger.error(f"Erro geral no processamento: {str(e)}")
            return False


# ============================================================================
# EXECUÇÃO
# ============================================================================

def main():
    """Função principal."""

    print("\n" + "="*70)
    print("  TestForge - Gerador Inteligente de Plano de Testes")
    print("  Engenheiro de QA Sênior & Especialista em Automação")
    print("="*70 + "\n")

    success = TestForgeController.process_documents()

    if success:
        print(f"\n✓ Sucesso! Arquivo '{OUTPUT_FILE}' gerado com sucesso!")
    else:
        print(f"\n✗ Erro ao gerar o plano de testes. Verifique o arquivo de log.")
        exit(1)


if __name__ == "__main__":
    main()